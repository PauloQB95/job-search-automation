"""Create the Excel workbook and static website dataset from scraper results."""

from datetime import date, datetime, timezone
import json
from pathlib import Path
import re

from openpyxl import load_workbook
import pandas as pd

from job_schema import JOB_COLUMNS

DATE_COLUMNS = ("Posting Date", "Closing Date")
EXCEL_DATE_FORMAT = "DD-MMM-YYYY"
WEBSITE_DATA_PATH = Path("docs") / "job_data.json"


def parse_job_date(value, source_name=""):
    """Return a date when a source value can be safely interpreted."""

    if value is None or pd.isna(value):
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    date_text = str(value).strip()

    if not date_text or date_text == "-":
        return None

    date_text = re.sub(r"\([^)]*\)$", "", date_text).strip()

    try:
        return datetime.fromisoformat(
            date_text.replace("Z", "+00:00")
        ).date()
    except ValueError:
        pass

    numeric_date_match = re.fullmatch(
        r"(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})",
        date_text
    )

    if numeric_date_match:
        first_number = int(numeric_date_match.group(1))
        second_number = int(numeric_date_match.group(2))

        if first_number > 12:
            numeric_format = "%d/%m/%Y" if "/" in date_text else "%d-%m-%Y"
        elif second_number > 12:
            numeric_format = "%m/%d/%Y" if "/" in date_text else "%m-%d-%Y"
        elif source_name in (
            "United Nations System",
            "United Nations System website"
        ):
            numeric_format = "%d/%m/%Y" if "/" in date_text else "%d-%m-%Y"
        elif source_name in ("API", "API - IDB", "API - World Bank"):
            numeric_format = "%m/%d/%Y" if "/" in date_text else "%m-%d-%Y"
        else:
            return None

        if len(numeric_date_match.group(3)) == 2:
            numeric_format = numeric_format.replace("%Y", "%y")

        return datetime.strptime(date_text, numeric_format).date()

    for date_format in (
        "%d %b %Y",
        "%d %B %Y",
        "%d %b %Y %I:%M %p",
        "%d %B %Y %I:%M %p",
        "%b %d, %Y",
        "%B %d, %Y",
        "%b %d, %Y, %I:%M:%S %p",
        "%B %d, %Y, %I:%M:%S %p",
        "%d/%b/%Y",
        "%d-%b-%Y",
        "%b-%d-%y",
        "%d %B %Y %H:%M",
        "%d %b %Y %H:%M",
        "%m/%d/%Y, %I:%M %p",
    ):
        try:
            return datetime.strptime(date_text, date_format).date()
        except ValueError:
            continue

    return None


def write_excel_workbook(jobs_dataframe, temporary_path):
    """Write a formatted workbook to a temporary path."""

    jobs_dataframe.to_excel(temporary_path, index=False)

    workbook = load_workbook(temporary_path)
    worksheet = workbook.active
    date_column_indexes = {
        column_name: JOB_COLUMNS.index(column_name) + 1
        for column_name in DATE_COLUMNS
    }
    source_column_index = JOB_COLUMNS.index("Source") + 1

    for column_index in date_column_indexes.values():
        for row_index in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            source_name = worksheet.cell(
                row=row_index,
                column=source_column_index
            ).value or ""
            parsed_date = parse_job_date(cell.value, source_name)

            if parsed_date is not None:
                cell.value = parsed_date
                cell.number_format = EXCEL_DATE_FORMAT
            elif cell.value == "-":
                cell.value = None

    workbook.save(temporary_path)


def write_website_data(scraper_results, temporary_path):
    """Write static website data from the results of a successful run."""

    website_jobs = []

    for source_name, jobs in scraper_results.items():
        for job in jobs:
            website_job = {
                column_name: job.get(column_name, "")
                for column_name in JOB_COLUMNS
            }

            for date_column in DATE_COLUMNS:
                parsed_date = parse_job_date(
                    website_job[date_column],
                    website_job["Source"]
                )

                if parsed_date is not None:
                    website_job[date_column] = parsed_date.isoformat()
                elif website_job[date_column] == "-":
                    website_job[date_column] = ""

            website_job["Dataset Source"] = source_name
            website_jobs.append(website_job)

    payload = {
        "generatedAt": datetime.now(timezone.utc).replace(
            microsecond=0
        ).isoformat().replace("+00:00", "Z"),
        "jobs": website_jobs,
    }

    with temporary_path.open("w", encoding="utf-8") as data_file:
        json.dump(payload, data_file, ensure_ascii=False, separators=(",", ":"))
